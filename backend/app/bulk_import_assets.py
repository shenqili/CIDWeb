from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import text

from .config import get_settings
from .db import SessionLocal

PROJECT_ID = "33333333-3333-3333-3333-333333333333"
DEMO_USER_ID = "22222222-2222-2222-2222-222222222222"
METRIC_CATALOG_VERSION_ID = "88888888-8888-8888-8888-888888888888"

BRAND_NAME_TO_CODE = {
    "BDF": "brand-bdf",
    "娇韵诗": "brand-jiaoyunshi",
    "丸美": "brand-wanmei",
    "溪木源": "brand-ximuyuan",
    "雅诗兰黛": "brand-estee",
    "云南白药": "brand-yunnanbaiyao",
    "拉芳": "brand-lafang",
    "汉高": "brand-hangao",
    "环亚": "brand-huanya",
}


@dataclass
class BrandContext:
    brand_id: str
    brand_code: str
    brand_name: str
    version_bundle_id: str


def normalize_field_name(source_name: str) -> str:
    normalized = re.sub(r"[^0-9A-Za-z]+", "_", source_name.strip()).strip("_").lower()
    return normalized or "field"


def metric_code_for(source_name: str) -> str:
    digest = hashlib.md5(source_name.encode("utf-8")).hexdigest()[:12]
    return f"metric_{digest}"


def instrument_family_for(source_name: str) -> str:
    if "_" in source_name:
        return source_name.split("_", 1)[0]
    return "derived"


def body_site_for(source_name: str) -> str | None:
    if "_" in source_name:
        return source_name.rsplit("_", 1)[-1]
    return None


def load_brand_contexts() -> dict[str, BrandContext]:
    with SessionLocal() as db:
        rows = db.execute(
            text(
                """
                select
                  b.brand_id,
                  b.brand_code,
                  b.brand_name,
                  pbb.default_version_bundle_id
                from brand b
                join project_brand_binding pbb on pbb.brand_id = b.brand_id
                where pbb.project_id = :project_id
                order by b.brand_code
                """
            ),
            {"project_id": PROJECT_ID},
        ).mappings().all()
    return {
        row["brand_code"]: BrandContext(
            brand_id=str(row["brand_id"]),
            brand_code=row["brand_code"],
            brand_name=row["brand_name"],
            version_bundle_id=str(row["default_version_bundle_id"]),
        )
        for row in rows
    }


def parse_metric_workbook(metric_workbook_path: Path) -> dict[str, set[str]]:
    workbook = load_workbook(metric_workbook_path, read_only=True, data_only=True)
    brand_metrics: dict[str, set[str]] = defaultdict(set)

    for worksheet in workbook.worksheets[:2]:
        rows = worksheet.iter_rows(values_only=True)
        _body_sites = [str(value).strip() if value is not None else "" for value in next(rows)]
        brand_names = [str(value).strip() if value is not None else "" for value in next(rows)]

        for row in rows:
            values = [str(value).strip() if value is not None else "" for value in row]
            for index, metric_name in enumerate(values):
                brand_name = brand_names[index] if index < len(brand_names) else ""
                brand_code = BRAND_NAME_TO_CODE.get(brand_name)
                if brand_code and metric_name:
                    brand_metrics[brand_code].add(metric_name)

    worksheet = workbook.worksheets[2]
    rows = list(worksheet.iter_rows(values_only=True))
    for row in rows[1:]:
        values = [str(value).strip() if value is not None else "" for value in row]
        brand_code = BRAND_NAME_TO_CODE.get(values[1] if len(values) > 1 else "")
        for site in values[2:]:
            if not brand_code or not site:
                continue
            for prefix in ["CM26DG_L*(D65)", "CM26DG_a*(D65)", "CM26DG_b*(D65)", "CM26DG_ITA", "CM26DG_GU"]:
                brand_metrics[brand_code].add(f"{prefix}_{site}")

    return brand_metrics


def parse_all_data_columns(all_data_path: Path, brand_metrics: dict[str, set[str]]) -> tuple[list[str], dict[str, list[int]]]:
    workbook = load_workbook(all_data_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value).strip() if value is not None else "" for value in next(worksheet.iter_rows(values_only=True))]

    metric_union = set()
    for metrics in brand_metrics.values():
        metric_union.update(metrics)

    common_columns = header[:10]
    unmatched_shared_columns = [column for column in header[10:] if column not in metric_union]
    shared_columns = set(common_columns + unmatched_shared_columns)

    brand_column_indexes: dict[str, list[int]] = {}
    for brand_code, metrics in brand_metrics.items():
        indexes = []
        for index, column in enumerate(header):
            if column in shared_columns or column in metrics:
                indexes.append(index)
        brand_column_indexes[brand_code] = indexes

    return header, brand_column_indexes


def ensure_metric_aliases(headers: list[str], brand_metrics: dict[str, set[str]]) -> None:
    unique_metric_fields = sorted({field for metrics in brand_metrics.values() for field in metrics if field in headers})

    with SessionLocal.begin() as db:
        for raw_field_name in unique_metric_fields:
            exists = db.execute(
                text(
                    """
                    select 1
                    from metric_alias
                    where metric_catalog_version_id = :metric_catalog_version_id
                      and raw_field_name = :raw_field_name
                    limit 1
                    """
                ),
                {
                    "metric_catalog_version_id": METRIC_CATALOG_VERSION_ID,
                    "raw_field_name": raw_field_name,
                },
            ).scalar_one_or_none()
            if exists:
                continue

            db.execute(
                text(
                    """
                    insert into metric_alias (
                      metric_alias_id,
                      metric_catalog_version_id,
                      raw_field_name,
                      standard_metric_code,
                      instrument_family,
                      body_site,
                      variable_type,
                      is_analyzable,
                      is_default_visible
                    )
                    values (
                      :metric_alias_id,
                      :metric_catalog_version_id,
                      :raw_field_name,
                      :standard_metric_code,
                      :instrument_family,
                      :body_site,
                      'number',
                      true,
                      true
                    )
                    """
                ),
                {
                    "metric_alias_id": str(uuid4()),
                    "metric_catalog_version_id": METRIC_CATALOG_VERSION_ID,
                    "raw_field_name": raw_field_name,
                    "standard_metric_code": metric_code_for(raw_field_name),
                    "instrument_family": instrument_family_for(raw_field_name),
                    "body_site": body_site_for(raw_field_name),
                },
            )


def ensure_brand_metric_mappings(brand_contexts: dict[str, BrandContext], brand_metrics: dict[str, set[str]]) -> None:
    with SessionLocal.begin() as db:
        db.execute(
            text(
                """
                create table if not exists brand_metric_mapping (
                  brand_metric_mapping_id uuid primary key default gen_random_uuid(),
                  brand_id uuid not null references brand(brand_id) on delete cascade,
                  metric_catalog_version_id uuid not null references metric_catalog_version(metric_catalog_version_id) on delete cascade,
                  raw_field_name varchar(255) not null,
                  instrument_family varchar(128),
                  body_site varchar(64),
                  source_sheet varchar(128),
                  created_at timestamptz not null default now(),
                  unique(brand_id, metric_catalog_version_id, raw_field_name)
                )
                """
            )
        )

        for brand_code, metrics in brand_metrics.items():
            brand = brand_contexts[brand_code]
            for raw_field_name in sorted(metrics):
                db.execute(
                    text(
                        """
                        insert into brand_metric_mapping (
                          brand_metric_mapping_id,
                          brand_id,
                          metric_catalog_version_id,
                          raw_field_name,
                          instrument_family,
                          body_site,
                          source_sheet
                        )
                        values (
                          :brand_metric_mapping_id,
                          :brand_id,
                          :metric_catalog_version_id,
                          :raw_field_name,
                          :instrument_family,
                          :body_site,
                          '各家指标(1).xlsx'
                        )
                        on conflict (brand_id, metric_catalog_version_id, raw_field_name) do update
                        set instrument_family = excluded.instrument_family,
                            body_site = excluded.body_site,
                            source_sheet = excluded.source_sheet
                        """
                    ),
                    {
                        "brand_metric_mapping_id": str(uuid4()),
                        "brand_id": brand.brand_id,
                        "metric_catalog_version_id": METRIC_CATALOG_VERSION_ID,
                        "raw_field_name": raw_field_name,
                        "instrument_family": instrument_family_for(raw_field_name),
                        "body_site": body_site_for(raw_field_name),
                    },
                )


def ensure_subjects(rd_values: list[str], brand_context: BrandContext) -> tuple[int, int]:
    created_subject_count = 0
    reused_subject_count = 0

    with SessionLocal.begin() as db:
        existing = {
            row["identifier_value"]: str(row["subject_id"])
            for row in db.execute(
                text(
                    """
                    select identifier_value, subject_id
                    from subject_identifier
                    where project_id = :project_id
                      and identifier_type = 'rd'
                      and identifier_value = any(:identifier_values)
                    """
                ),
                {
                    "project_id": PROJECT_ID,
                    "identifier_values": rd_values,
                },
            ).mappings().all()
        }

        for rd_value in rd_values:
            if rd_value in existing:
                reused_subject_count += 1
                continue

            subject_id = str(uuid4())
            db.execute(
                text(
                    """
                    insert into subject (subject_id, project_id, brand_id, subject_status)
                    values (:subject_id, :project_id, :brand_id, 'active')
                    """
                ),
                {
                    "subject_id": subject_id,
                    "project_id": PROJECT_ID,
                    "brand_id": brand_context.brand_id,
                },
            )
            db.execute(
                text(
                    """
                    insert into subject_identifier (
                      subject_identifier_id,
                      subject_id,
                      project_id,
                      identifier_type,
                      identifier_value,
                      is_primary,
                      source,
                      created_by
                    )
                    values (
                      :subject_identifier_id,
                      :subject_id,
                      :project_id,
                      'rd',
                      :identifier_value,
                      true,
                      'bulk_import',
                      :created_by
                    )
                    """
                ),
                {
                    "subject_identifier_id": str(uuid4()),
                    "subject_id": subject_id,
                    "project_id": PROJECT_ID,
                    "identifier_value": rd_value,
                    "created_by": DEMO_USER_ID,
                },
            )
            existing[rd_value] = subject_id
            created_subject_count += 1

    return created_subject_count, reused_subject_count


def get_or_create_dataset(brand_context: BrandContext, dataset_name: str) -> str:
    with SessionLocal.begin() as db:
        existing = db.execute(
            text(
                """
                select dataset_id
                from dataset
                where project_id = :project_id
                  and brand_id = :brand_id
                  and dataset_name = :dataset_name
                limit 1
                """
            ),
            {
                "project_id": PROJECT_ID,
                "brand_id": brand_context.brand_id,
                "dataset_name": dataset_name,
            },
        ).scalar_one_or_none()

        if existing is not None:
            return str(existing)

        dataset_id = str(uuid4())
        db.execute(
            text(
                """
                insert into dataset (dataset_id, project_id, brand_id, dataset_name, dataset_type, status)
                values (:dataset_id, :project_id, :brand_id, :dataset_name, 'instrument', 'active')
                """
            ),
            {
                "dataset_id": dataset_id,
                "project_id": PROJECT_ID,
                "brand_id": brand_context.brand_id,
                "dataset_name": dataset_name,
            },
        )
        return dataset_id


def dataset_already_published(dataset_id: str, source_file_name: str) -> bool:
    with SessionLocal() as db:
        existing = db.execute(
            text(
                """
                select 1
                from dataset_version dv
                join import_batch ib on ib.import_batch_id = dv.source_import_batch_id
                where dv.dataset_id = :dataset_id
                  and ib.source_file_name = :source_file_name
                  and dv.status = 'published'
                limit 1
                """
            ),
            {
                "dataset_id": dataset_id,
                "source_file_name": source_file_name,
            },
        ).scalar_one_or_none()
    return existing is not None


def next_dataset_version(dataset_id: str) -> tuple[str, str | None]:
    with SessionLocal() as db:
        rows = db.execute(
            text(
                """
                select dataset_version_id, version_no
                from dataset_version
                where dataset_id = :dataset_id
                order by created_at
                """
            ),
            {"dataset_id": dataset_id},
        ).mappings().all()

    if not rows:
        return "v1", None

    last = rows[-1]
    last_no = str(last["version_no"]).lstrip("v")
    next_no = int(last_no) + 1 if last_no.isdigit() else len(rows) + 1
    return f"v{next_no}", str(last["dataset_version_id"])


def write_filtered_csv(
    all_data_path: Path,
    selected_indexes: list[int],
    output_path: Path,
) -> tuple[list[str], int, list[str]]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(all_data_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)

    header = [str(value).strip() if value is not None else "" for value in next(rows)]
    selected_header = [header[index] for index in selected_indexes]
    rd_index = selected_header.index("编号：RD")
    rd_values: list[str] = []
    row_count = 0

    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(selected_header)

        for row in rows:
            values = [row[index] if index < len(row) else None for index in selected_indexes]
            serialized = ["" if value is None else value for value in values]
            rd_value = str(serialized[rd_index]).strip()
            if not rd_value:
                continue
            writer.writerow(serialized)
            rd_values.append(rd_value)
            row_count += 1

    return selected_header, row_count, rd_values


def save_manifest(settings, import_batch_id: str, manifest: dict) -> None:
    manifest_dir = Path(settings.manifest_storage_root)
    manifest_dir.mkdir(parents=True, exist_ok=True)
    (manifest_dir / f"{import_batch_id}.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def publish_brand_dataset(
    settings,
    brand_context: BrandContext,
    selected_headers: list[str],
    row_count: int,
    rd_values: list[str],
    source_file_name: str,
    storage_path: str,
    dataset_id: str,
) -> None:
    created_subject_count, reused_subject_count = ensure_subjects(rd_values, brand_context)
    version_no, parent_version_id = next_dataset_version(dataset_id)
    import_batch_id = str(uuid4())
    dataset_version_id = str(uuid4())
    now = datetime.now(timezone.utc)
    dataset_hash = hashlib.md5(f"{brand_context.brand_code}:{source_file_name}:{row_count}:{len(selected_headers)}".encode("utf-8")).hexdigest()

    with SessionLocal.begin() as db:
        db.execute(
            text(
                """
                insert into import_batch (
                  import_batch_id,
                  dataset_id,
                  brand_id,
                  version_bundle_id,
                  source_file_name,
                  storage_path,
                  import_mode,
                  status,
                  created_by
                )
                values (
                  :import_batch_id,
                  :dataset_id,
                  :brand_id,
                  :version_bundle_id,
                  :source_file_name,
                  :storage_path,
                  'append',
                  'published',
                  :created_by
                )
                """
            ),
            {
                "import_batch_id": import_batch_id,
                "dataset_id": dataset_id,
                "brand_id": brand_context.brand_id,
                "version_bundle_id": brand_context.version_bundle_id,
                "source_file_name": source_file_name,
                "storage_path": storage_path,
                "created_by": DEMO_USER_ID,
            },
        )

        db.execute(
            text(
                """
                insert into dataset_version (
                  dataset_version_id,
                  dataset_id,
                  version_no,
                  parent_version_id,
                  source_import_batch_id,
                  status,
                  is_derived,
                  row_count,
                  column_count,
                  dataset_hash,
                  created_by,
                  published_at
                )
                values (
                  :dataset_version_id,
                  :dataset_id,
                  :version_no,
                  :parent_version_id,
                  :source_import_batch_id,
                  'published',
                  false,
                  :row_count,
                  :column_count,
                  :dataset_hash,
                  :created_by,
                  :published_at
                )
                """
            ),
            {
                "dataset_version_id": dataset_version_id,
                "dataset_id": dataset_id,
                "version_no": version_no,
                "parent_version_id": parent_version_id,
                "source_import_batch_id": import_batch_id,
                "row_count": row_count,
                "column_count": len(selected_headers),
                "dataset_hash": dataset_hash,
                "created_by": DEMO_USER_ID,
                "published_at": now,
            },
        )

        if parent_version_id:
            db.execute(
                text(
                    """
                    insert into dataset_version_lineage (
                      dataset_version_lineage_id,
                      from_dataset_version_id,
                      to_dataset_version_id,
                      lineage_type
                    )
                    values (
                      :dataset_version_lineage_id,
                      :from_dataset_version_id,
                      :to_dataset_version_id,
                      'bulk_import'
                    )
                    """
                ),
                {
                    "dataset_version_lineage_id": str(uuid4()),
                    "from_dataset_version_id": parent_version_id,
                    "to_dataset_version_id": dataset_version_id,
                },
            )

        subject_rows = {
            row["identifier_value"]: str(row["subject_id"])
            for row in db.execute(
                text(
                    """
                    select identifier_value, subject_id
                    from subject_identifier
                    where project_id = :project_id
                      and identifier_type = 'rd'
                      and identifier_value = any(:identifier_values)
                    """
                ),
                {
                    "project_id": PROJECT_ID,
                    "identifier_values": rd_values,
                },
            ).mappings().all()
        }

        for sequence, rd_value in enumerate(rd_values, start=1):
            db.execute(
                text(
                    """
                    insert into visit_record (
                      visit_id,
                      subject_id,
                      dataset_version_id,
                      import_batch_id,
                      system_batch_seq,
                      visit_label
                    )
                    values (
                      :visit_id,
                      :subject_id,
                      :dataset_version_id,
                      :import_batch_id,
                      :system_batch_seq,
                      :visit_label
                    )
                    """
                ),
                {
                    "visit_id": str(uuid4()),
                    "subject_id": subject_rows[rd_value],
                    "dataset_version_id": dataset_version_id,
                    "import_batch_id": import_batch_id,
                    "system_batch_seq": sequence,
                    "visit_label": rd_value,
                },
            )

    manifest = {
        "importBatchId": import_batch_id,
        "datasetId": dataset_id,
        "sourceFileName": source_file_name,
        "fileName": source_file_name,
        "rowCount": row_count,
        "columnCount": len(selected_headers),
        "detectedColumns": selected_headers,
        "rows": [],
        "fieldMappings": [{"sourceName": column, "targetName": normalize_field_name(column)} for column in selected_headers],
        "fieldMappingStatus": "completed",
        "primaryKeyColumn": "编号：RD",
        "identifierType": "rd",
        "primaryKeyMappingStatus": "completed",
        "createdSubjectCount": created_subject_count,
        "reusedSubjectCount": reused_subject_count,
        "missingPrimaryKeyRows": [],
        "publishedDatasetVersionId": dataset_version_id,
        "publishedVersionNo": version_no,
    }
    save_manifest(settings, import_batch_id, manifest)


def main() -> None:
    settings = get_settings()
    storage_root = Path(settings.raw_storage_root)
    candidate_roots = [
        Path(__file__).resolve().parents[2] / "docs",
        Path(__file__).resolve().parents[1] / "import_source",
        Path("/storage/import_source"),
    ]
    source_root = next((root for root in candidate_roots if root.exists()), None)
    if source_root is None:
        raise FileNotFoundError("No import source directory found. Expected docs, backend/import_source, or /storage/import_source.")

    all_data_path = source_root / "alldata_260204(1).xlsx"
    metric_workbook_path = next(path for path in source_root.glob("*.xlsx") if path.name != "alldata_260204(1).xlsx")

    brand_contexts = load_brand_contexts()
    brand_metrics = parse_metric_workbook(metric_workbook_path)
    header, brand_column_indexes = parse_all_data_columns(all_data_path, brand_metrics)

    ensure_metric_aliases(header, brand_metrics)
    ensure_brand_metric_mappings(brand_contexts, brand_metrics)

    summary = []
    for brand_code, indexes in brand_column_indexes.items():
        brand_context = brand_contexts[brand_code]
        dataset_name = f"alldata_260204_full_{brand_code}"
        dataset_id = get_or_create_dataset(brand_context, dataset_name)
        source_file_name = f"alldata_260204_{brand_code}.csv"

        if dataset_already_published(dataset_id, source_file_name):
            summary.append({"brand": brand_code, "status": "skipped_existing", "datasetId": dataset_id})
            continue

        output_path = storage_root / dataset_id / source_file_name
        selected_headers, row_count, rd_values = write_filtered_csv(all_data_path, indexes, output_path)
        publish_brand_dataset(
            settings=settings,
            brand_context=brand_context,
            selected_headers=selected_headers,
            row_count=row_count,
            rd_values=rd_values,
            source_file_name=source_file_name,
            storage_path=str(output_path),
            dataset_id=dataset_id,
        )
        summary.append(
            {
                "brand": brand_code,
                "status": "imported",
                "datasetId": dataset_id,
                "rowCount": row_count,
                "columnCount": len(selected_headers),
            }
        )

    print(json.dumps({"projectId": PROJECT_ID, "summary": summary}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
