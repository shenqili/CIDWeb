import csv
import hashlib
import json
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from .config import get_settings


settings = get_settings()


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _coerce_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_uploaded_file(file_name: str, content: bytes) -> dict:
    suffix = Path(file_name).suffix.lower()
    if suffix in {".csv", ".txt"}:
        return _parse_delimited(content)
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_xlsx(content)
    raise ValueError(f"暂不支持的文件类型: {suffix}")


def _parse_delimited(content: bytes) -> dict:
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if decoded is None:
        raise ValueError("无法识别文件编码")

    sample = decoded[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(StringIO(decoded), dialect=dialect)
    columns = [str(column).strip() for column in (reader.fieldnames or []) if str(column).strip()]
    rows = []
    for raw_row in reader:
        if raw_row is None:
            continue
        rows.append({column: _coerce_cell(raw_row.get(column)) for column in columns})

    if not columns:
        raise ValueError("无法检测到文件表头")

    return {"columns": columns, "rows": rows}


def _parse_xlsx(content: bytes) -> dict:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration as exc:
        workbook.close()
        raise ValueError("Excel 文件为空") from exc

    columns = [_coerce_cell(value) for value in header_row if _coerce_cell(value)]
    if not columns:
        workbook.close()
        raise ValueError("无法检测到 Excel 表头")

    rows = []
    for row in iterator:
        if row is None:
            continue
        values = list(row)
        rows.append(
            {
                columns[index]: _coerce_cell(values[index]) if index < len(values) else ""
                for index in range(len(columns))
            }
        )

    workbook.close()
    return {"columns": columns, "rows": rows}


def build_manifest(import_batch_id: str, dataset_id: str, source_file_name: str, file_name: str, content: bytes) -> dict:
    parsed = parse_uploaded_file(file_name=file_name, content=content)
    return {
        "importBatchId": import_batch_id,
        "datasetId": dataset_id,
        "sourceFileName": source_file_name,
        "detectedColumns": parsed["columns"],
        "rows": parsed["rows"],
        "rowCount": len(parsed["rows"]),
        "columnCount": len(parsed["columns"]),
        "fieldMappings": [],
        "fieldMappingStatus": "pending",
        "primaryKeyColumn": None,
        "identifierType": None,
        "primaryKeyMappingStatus": "pending",
        "createdSubjectCount": 0,
        "reusedSubjectCount": 0,
        "missingPrimaryKeyRows": [],
        "publishedDatasetVersionId": None,
        "publishedVersionNo": None,
    }


def get_manifest_path(import_batch_id: str) -> Path:
    return Path(settings.manifest_storage_root) / f"{import_batch_id}.json"


def save_manifest(manifest: dict) -> Path:
    path = get_manifest_path(manifest["importBatchId"])
    _ensure_parent(path)
    path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def load_manifest(import_batch_id: str) -> dict:
    path = get_manifest_path(import_batch_id)
    if not path.exists():
        raise FileNotFoundError(f"未找到导入批次 manifest: {import_batch_id}")
    return json.loads(path.read_text(encoding="utf-8"))


def compute_rows_hash(rows: list[dict]) -> str:
    payload = json.dumps(rows, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def next_dataset_version_no(existing_versions: list[str]) -> str:
    numeric_suffixes = []
    for version in existing_versions:
        if version.startswith("v") and version[1:].isdigit():
            numeric_suffixes.append(int(version[1:]))
    next_index = max(numeric_suffixes, default=0) + 1
    return f"v{next_index}"


def next_brand_config_version_no(current_version: str | None) -> str:
    if not current_version:
        return "1.0.0"

    parts = current_version.split(".")
    if len(parts) != 3 or not all(part.isdigit() for part in parts):
        return "1.0.0"

    major, minor, patch = [int(part) for part in parts]
    return f"{major}.{minor}.{patch + 1}"
